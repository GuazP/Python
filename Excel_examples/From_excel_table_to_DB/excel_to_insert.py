import openpyxl
from string import ascii_uppercase

def main(args):
    print("Opening file")
    while True:
        try:
            if len(args) == 2:
                fil = args[1]
                wb = openpyxl.load_workbook(args[1])
                break
            elif len(args) == 1:
                fil = input("Nazwa pliku:")
                #~ fil = "example.xlsx"
                wb = openpyxl.load_workbook(fil)
                break
            else:
                raise ValueError
        except FileNotFoundError as ex:
            wb = openpyxl.load_workbook(input("Pliku nie znaleziono, wprowadź poprawnie nazwę:"))
        except ValueError as ex:
            print("Za dużo argumentów z linii komend.")
            exit(1);
    fil = "".join(txt for txt in fil.split(".")[:-1])
    print("Start")
    sheet = wb.active
    columns_length = ascii_uppercase[sheet.max_column -1]

    tables = []
    columns = []
    columns_values = []
    scrap = []

    start = 0
    stop = 1
    print("Collect first row")
    for rowOfCellObjects in sheet["A1":str(columns_length)+str(1)]:
        for cellObj in rowOfCellObjects:
            table, column = str(cellObj.value).split(".", 1)
            if not tables:
                tables.append(table)
            elif table in tables:
                stop += 1
            else:
                tables.append(table)
                scrap.append( (start, stop) )
                start = stop
                stop += 1
            columns.append(column)
    scrap.append( (start, stop) )

    print("Collect data")
    with open((fil[:]+".txt"), "w+") as f:
        for rowOfCellObjects in sheet["A2":str(columns_length)+str(sheet.max_row)]:
            for cellObj in rowOfCellObjects:
                columns_values.append(cellObj.value)

            for rang in scrap:
                if any(val!=None for val in columns_values[rang[0]:rang[1]]):
                    table_build = tables[scrap.index(rang)]
                    columns_build = [str(col).split(" ", 1)[0] for col in columns[rang[0]:rang[1]]]
                    values_build = [str(val) for val in columns_values[rang[0]:rang[1]]]
                    #Check txt
                    for col, val, itr in zip(columns[rang[0]:rang[1]], columns_values[rang[0]:rang[1]], range(len(columns_build))):
                        col = col.split(" ", 1)
                        if len(col) > 1:
                            if col[1] == "txt":
                                values_build[itr] = '"' + val + '"'
                            if col[1] == "int":
                                try:
                                    int(val)
                                except ValueError:
                                    exit("FAILED: Integer values, must contain only digits.")
                    f.write("INSERT INTO {}\n\t({})\nVALUES \n\t({});\n".format(table_build,
                                                                                ",\n\t".join(col for col in columns_build),\
                                                                                ",\n\t".join(val for val in values_build) ))
            f.write("\n")
            columns_values = []
    print("End")

if __name__=="__main__":
    import sys
    main(sys.argv)

